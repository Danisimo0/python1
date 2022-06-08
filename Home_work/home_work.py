import openpyxl
from bottle import get
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

book = Workbook()
stranica1 = book.active

var_range = range(1, 26)
line_1 = []
line1 = ""
for i in var_range:
    line1 = f"_{i}"
    line_1.append(line1)

line_2 = []
line2 = ""
for j in var_range:
    line2 = f"A_{j}"
    line_2.append(line2)

line_final = []
line_final.append(line_1)
line_final.append(line_2)
# print(stroke_final)

for row in range(0, len(line_final)):
    for col in range(0, len(line_final[row])):
        # print(stroka_final[row])
        stranica1[f"{get_column_letter(col + 1)}{row + 1}"] = line_final[row][col]

for col in range(0, len(line_final)):
    for row in range(0, len(line_final[col])):
        stranica1.cell(row=row + 4, column=col + 1).value = line_final[col][row]
print(line_final)

book.save("home_work_excel2.xlsx")
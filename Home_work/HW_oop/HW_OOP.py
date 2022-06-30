import math
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

filename = "HW.xlsx"
filedir = "C:\Project\projects\python1\python1\Home_work\HW_oop"
workbook = openpyxl.load_workbook(filedir + "/" + filename)
worksheet = workbook.active
max_row = worksheet.max_row
max_column = worksheet.max_column
product = []


class Produrty:
    def __init__(self, product: str, group: str, provider: str, data_postavki: float, region: str, seales: float,
                 sbyt: float, profit: str):
        self.tovar = product
        self.group = group
        self.postavshik = provider
        self.data_postavki = data_postavki
        self.region = region
        self.prodazhi = seales
        self.sbyt = sbyt
        self.pribyl = profit

    def zarobotok(self):
        return f"{self.tovar},{self.prodazhi},{self.pribyl}"

    def pribyl_dir(self):
        return f"{self.prodazhi},{self.pribyl}"


Global_masiv = []
for row in range(1, max_row + 1):
    tovar_add = []
    for col in range(1, max_column + 1):
        obj = worksheet.cell(row=row, column=col).value

        tovar_add.append(obj)

    name = Produrty(
        product=tovar_add[0],
        group=tovar_add[1],
        provider=tovar_add[2],
        data_postavki=tovar_add[3],
        region=tovar_add[4],
        seales=tovar_add[5],
        sbyt=tovar_add[6],
        profit=tovar_add[7]
    )

    obshee1 = name.zarobotok()
    pribl1 = name.pribyl_dir()
    obshee_masiv = obshee1.split(",")
    Global_masiv.append(obshee_masiv)
# print(thee_massive1)
book = Workbook()
sheet = book.active

for row in range(0, len(Global_masiv)):
    for col in range(0, len(Global_masiv[row])):
        sheet[f"{get_column_letter(col + 1)}{row + 1}"] = Global_masiv[row][col]
print(Global_masiv)

book.save("oop.xlsx")

print("\n\n\n*********************************\n\n\n")


class Generator:
    def __init__(self, name: str, index: str, value: str):
        self.name = name
        self.index = index
        self.value = value

    def get_name(self):
        return f"{self.name} {self.index} {self.value}"


ran = []
for i in range(1, 1001):
    ran.append(Generator(name=f"_{i}", index=f"{i}", value=f"_A{i}").get_name())
print(ran)

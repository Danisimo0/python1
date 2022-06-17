import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval
from openpyxl.utils import get_column_letter

first_file = "Folder 1.xlsx"
second_file = "Folder 2.xlsx"
third_file = "Folder 3.xlsx"
first_file_list_global = []
second_file_list_global = []
third_file_list_global = []
files = [[first_file, first_file_list_global], [second_file, second_file_list_global],
         [third_file, third_file_list_global]]
glob_list = []




# загружаем в память уже существующий файл на диске
for i in files:
    workbook = openpyxl.load_workbook(i[0])
    worksheet = workbook.active
    for y in range(1, 26):
        mass_local = []
        for x in range(1, 26):
            value = worksheet.cell(row=x, column=y).value
            if value is None:
                break
            mass_local.append(value)
        glob_list.append(mass_local)

    def iii():
        for ii in glob_list:
            # print("ii" + str(ii[0]))
            if len(ii) == 0:
                # print("ii")
                del glob_list[0]
                iii()

file_name = "task2.xlsx"

wb = Workbook()
# ws0 = wb.create_sheet("Folder 1")
ind_element = 1
for element in glob_list:
    if len(element) == 0:
        continue
    else:
        ws = wb.create_sheet("Лист" + str(ind_element))
        row = 0
        col = 1
        while True:
            row += 1
            ws.cell(row=row, column=col).value = str(element[row-1])
            if row == len(element):
                break

        ind_element += 1

if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

wb.save("HW.excel_2.xlsx")